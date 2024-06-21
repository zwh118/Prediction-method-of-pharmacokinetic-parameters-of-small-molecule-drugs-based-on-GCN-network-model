# -*- coding: utf-8 -*-
"""
@author: wangying

@date:2021.09.18
"""
## Calculate inter-molecular similarity

import xlrd
from rdkit import Chem
from rdkit import DataStructs
from pprint import pprint
import openpyxl

# input smiles
file = xlrd.open_workbook(r'F:\WY\OBA\oba_0.25.xlsx')
ws = file.sheet_by_name('valid')
datasets = []
c = 1
for r in range(1, ws.nrows):
    col = []
    datasets.append(ws.cell(r, c).value)

# Read in smiles and convert them into molecular objects, establish a list of moles
l = len(datasets)
print(l)
mols = []
for dataset in datasets:
    m = Chem.MolFromSmiles(dataset)
    mols.append(m)

# Calculate molecular fingerprints and compare molecular similarities

fps = [Chem.RDKFingerprint(x) for x in mols]  # Topological fingerprint
pprint(len(fps))


ppbr = []
for i in range(0, l):
    #print(i)
    mol1 = fps[i]
    smis = []
    sum = 0
    for j in range(0, l):
        #print(j)
        mol2 = fps[j]
        smi = DataStructs.FingerprintSimilarity(mol1, mol2)
        sum = sum + smi
        #print(smi)

        smis.append(smi)

    pprint(len(smis))
    average_smi = (sum - 1) / (l - 1)
    print(average_smi)
    print("average similarity between mol and mols: %.2f" %average_smi)

    # Filter data

    if average_smi >= 0.3:
        ppbr.append(datasets[i])
print(len(ppbr))

#Write to Excel
wb = openpyxl.Workbook()
ws = wb.create_sheet('Sheet1')

l1 = len(ppbr)
for r1 in range(0, l1):

    ws.cell(row=r1+1, column=1).value = ppbr[r1]
#print(len(ppbr[r1]))

wb.save(r'C:\Users\Desktop\oba(0.3).xlsx')
