# -*- coding: utf-8 -*-
"""
@author: wangying

@date:2022.11.28
"""


from rdkit import Chem
import numpy as np
import scipy.sparse as sp
from rdkit.Chem import Draw
from rdkit.Chem.rdmolops import GetAdjacencyMatrix
import torch.nn as nn
import torch.nn.functional as F
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import r2_score # R square
from math import sqrt
#from pygcn.layers import GraphConvolution
import deepchem as dc
import torch

from torch.utils.data import Dataset, DataLoader

import xlrd
import openpyxl
import os

import math
from torch.nn.parameter import Parameter
from torch.nn.modules.module import Module
from sklearn import linear_model
import joblib

max_natoms = 457
EPOCH = 30

#train_num = 299
#valid_num = 100

train_num = 264
valid_num = 90

#train_num = 1166
#valid_num = 389

train_batchsize = 66
valid_batchsize = 30

#train_batchsize = 73
#valid_batchsize = 24

if (train_num % train_batchsize == 0):
    q1 = int(train_num / train_batchsize)
else:
    q1 = int(train_num / train_batchsize) + 1

if (valid_num % valid_batchsize == 0):
    q2 = int(valid_num / valid_batchsize)
else:
    q2 = int(valid_num / valid_batchsize) + 1
print(q1)
print(q2)

file = xlrd.open_workbook(r'F:\WY\PPBR\ppbr(0.25).xlsx')


ws = file.sheet_by_name('train')
ws1 = file.sheet_by_name('valid')


Y = []
S = []
for r in range(1, train_num+1):
    y = ws.cell(r, 2).value
    y = [y]
    Y.append(y)

    s = ws.cell(r, 1).value
    s = s
    S.append(s)


Y1 = []
S1 = []
for r in range(1, valid_num+1):
    y = ws1.cell(r, 2).value
    y = [y]
    Y1.append(y)

    s = ws1.cell(r, 1).value
    s = s
    S1.append(s)

file.release_resources()
del file



class MolDataset(Dataset):
    def __init__(self, smiles, properties, max_natoms, normalize_A=False):
        self.smiles = smiles
        self.properties = properties
        self.max_natoms = max_natoms

    def __len__(self):
        return len(self.smiles)

    def __getitem__(self, idx):
        s = self.smiles[idx]
        m = Chem.MolFromSmiles(s)
        natoms = m.GetNumAtoms()
        #print(idx)
        #print(s)
        #print(natoms)

        #A = GetAdjacencyMatrix(m) + np.eye(natoms)
        A = GetAdjacencyMatrix(m)
        #print('A:', A)

        A_padding = np.zeros((self.max_natoms, self.max_natoms))
        A_padding[:natoms, :natoms] = A



        I = np.eye(A_padding.shape[0])
        A_hat = A_padding + I

        D_hat = np.array(np.sum(A_hat, axis=0))

        D_hat = np.power(D_hat, -0.5)  # !!!!
        D_hat = np.diag(D_hat)  # A_hat
        #D_hat1 = np.linalg.inv(D_hat)
        #A = np.dot(D_hat1, A_hat)
        A1 = np.dot(D_hat, A_hat)
        A = np.dot(A1, D_hat)





        X = [self.atom_feature(m, i) for i in range(natoms)]
        for i in range(natoms, max_natoms):
            X.append(np.zeros(78))
        X = np.array(X)
        #print('X_:', X)


        sample = dict()
        sample['X'] = torch.from_numpy(X)
        sample['A'] = torch.from_numpy(A)
        sample['Y'] = torch.from_numpy(np.array(self.properties[idx]))
        #print(sample)

        return sample



    # one-hot
    def one_of_k_encoding(self, x, allowable_set):
        if x not in allowable_set:
            raise Exception("input {0} not in allowable set{1}:".format(x, allowable_set))
        # print list((map(lambda s: x == s, allowable_set)))
        return list(map(lambda s: x == s, allowable_set))

    def one_of_k_encoding_unk(self, x, allowable_set):
        """Maps inputs not in the allowable set to the last element."""
        if x not in allowable_set:
            x = allowable_set[-1]
        return list(map(lambda s: x == s, allowable_set))

    def atom_feature(self, m, atom_i):

        atom = m.GetAtomWithIdx(atom_i)

        return np.array((self.one_of_k_encoding_unk(atom.GetSymbol(),
                                               ['C', 'N', 'O', 'S', 'F', 'Si', 'P', 'Cl', 'Br', 'Mg', 'Na',
                                                'Ca', 'Fe', 'As', 'Al', 'I', 'B', 'V', 'K', 'Tl', 'Yb',
                                                'Sb', 'Sn', 'Ag', 'Pd', 'Co', 'Se', 'Ti', 'Zn', 'H',  # H?
                                                'Li', 'Ge', 'Cu', 'Au', 'Ni', 'Cd', 'In', 'Mn', 'Zr',
                                                'Cr', 'Pt', 'Hg', 'Pb', 'other']) +  # 'C', 'O', 'N', 'S', 'Cl', 'F', 'Br', 'P', 'I', 'Si', 'B', 'Na', 'Sn', 'Se', 'other'
                         self.one_of_k_encoding(atom.GetDegree(), [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) +
                         self.one_of_k_encoding_unk(atom.GetTotalNumHs(), [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) +
                         self.one_of_k_encoding_unk(atom.GetImplicitValence(), [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) +
                         [atom.GetIsAromatic()]), dtype=float)  # (15, 7, 5, 7, 1) --> total 28 , (44, 7, 5, 7, 1) --> total 28


class GraphConvolution(Module):
    def __init__(self, in_features, out_features, bias=True):
        super(GraphConvolution, self).__init__()
        self.in_features = in_features
        self.out_features = out_features
        self.weight = Parameter(torch.FloatTensor(in_features, out_features))
        if bias:
            self.bias = Parameter(torch.FloatTensor(out_features))
        else:
            self.register_parameter('bias', None)
        self.reset_parameters()

    def reset_parameters(self):
        stdv = 1. / math.sqrt(self.weight.size(1))
        self.weight.data.uniform_(-stdv, stdv)
        if self.bias is not None:
            self.bias.data.uniform_(-stdv, stdv)

    def forward(self, input, adj):
        #support = torch.mm(input, self.weight)
        support = torch.matmul(input, self.weight)

        #output = torch.spmm(adj, support)
        output = torch.matmul(adj, support)
        if self.bias is not None:
            return output + self.bias
        else:
            return output

    def __repr__(self):
        return self.__class__.__name__ + ' (' \
               + str(self.in_features) + ' -> ' \
               + str(self.out_features) + ')'


class GCN(nn.Module):
    def __init__(self, nfeat, nhid, dropout):

        super(GCN, self).__init__()

        self.embedding = nn.Linear(78, nfeat)
        self.gc1 = GraphConvolution(nfeat, 50)
        self.gc2 = GraphConvolution(50, 30)
        self.gc3 = GraphConvolution(30, nhid)
        self.fc = nn.Linear(nhid, 1)
        self.dropout = dropout
    def forward(self, x, adj):
        x = self.embedding(x)
        #x = torch.einsum('ijk,ikl->ijl', (A, x))
        x = F.relu(self.gc1(x, adj))
        x = F.relu(self.gc2(x, adj))
        x = self.gc3(x, adj)
        x = F.dropout(x, self.dropout, training=self.training)
        x = F.relu(x)
       # x = torch.sum(x, dim=1)
        x = x.mean(1)
        #retval = self.fc(x)
        #retval = F.softmax(retval, dim=1)

        # x = F.relu(x)
        retval = self.fc(x)
        return retval
        # return F.log_softmax(retval, dim=1)


lr = 0.0001


gcn_model = GCN(80, 10, 0.2).cuda()
mlr = linear_model.LinearRegression()



for param in gcn_model.parameters():
    if param.dim() == 1:
        continue
        nn.init.constant(param, 0)
    else:
        nn.init.xavier_normal_(param)



train_dataset = MolDataset(S, Y, max_natoms)
valid_dataset = MolDataset(S1, Y1, max_natoms)


train_dataloader = DataLoader(train_dataset, batch_size=train_batchsize, num_workers=0, shuffle=True)
valid_dataloader = DataLoader(valid_dataset, batch_size=valid_batchsize, num_workers=0, shuffle=True)

optimizer = torch.optim.Adam(gcn_model.parameters(), lr=lr, weight_decay = 100.0)
#optimizer = torch.optim.SGD(model.parameters(), lr=lr)

loss_fn = nn.MSELoss()
loss_list = []
epoch_acc = 0
MAE = []
RMSE = []
MAE1 = []
RMSE1 = []

for epoch in range(EPOCH):
  epoch_loss = []

  print('epoch:', epoch)
  accuracy2 = 0
  best_model = None
  val_bestacc = 0

  VAL_MAE = 0
  VAL_RMSE = 0

  for i_batch, batch in enumerate(train_dataloader):
      #print(i_batch)

      x, y, A = \
        batch['X'].cuda().float(), batch['Y'].cuda().float(), batch['A'].cuda().float()

      pred = gcn_model(x, A).squeeze(-1)

      y1 = y.cpu().numpy()


      #print('Y:', y1)

      pred1 = pred.detach().cpu().numpy()


      RF_X = []
      for i in range(len(pred1)):
          pred2 = [pred1[i]]
          RF_X.append(pred2)

      #print("trian_y:", y1)

      mlr.fit(RF_X, y1)

      RF_train_predict = mlr.predict(RF_X)

      RF_train_Pre = []
      for i in range(len(RF_train_predict)):
          Pred2 = RF_train_predict[i]
          RF_train_Pre.append(Pred2)

      #print('train_pre:', RF_train_Pre)



      a = []
      b = abs(RF_train_Pre - y1)
      #print(b)
      #print(len(b))

      for i in range(len(b)):
          if b[i] <= 0.1:  #!!!!!!!
               a.append(b[i])

      l = len(a)
       # print(len(a))

      accuracy1 = float(l) / len(b)
      accuracy1 = ('%.2f' % accuracy1)
      #print(accuracy1)

      accuracy2 = accuracy2 + float(accuracy1)
      Accuracy = accuracy2 / q1



      mae = mean_absolute_error(y1, RF_train_Pre)
      rmse = sqrt(mean_squared_error(y1, RF_train_Pre))
      #R2 = r2_score(y1, Pred)
      MAE.append(mae)
      RMSE.append(rmse)

      '''
      # 计算loss
      loss = loss_fn(pred, y.squeeze(-1))
      loss.backward()
      torch.nn.utils.clip_grad_norm_(XG_model.parameters(), 1.0)

      optimizer.step()
      loss_list.append(loss.data.cpu().numpy())
      epoch_loss.append(loss.data.cpu().numpy())
      '''


      val_accuracy2 = 0
      i_VAL_MAE = 0
      i_VAL_RMSE = 0
      for i_batch, batch in enumerate(valid_dataloader):
          # print('val',i_batch)
          x, y, A = \
              batch['X'].cuda().float(), batch['Y'].cuda().float(), batch['A'].cuda().float()

          val_y1 = y.cpu().numpy()
          val_pred = gcn_model(x, A).squeeze(-1)

          val_pred1 = val_pred.detach().cpu().numpy()
          RF_val_x = []
          for i in range(len(val_pred1)):
              val_pred2 = [val_pred1[i]]
              RF_val_x.append(val_pred2)

          #print('val_y:', val_y1)
          #XG_model.fit(XG_val_x, val_y1)
          mlr.fit(RF_val_x, val_y1)

          RF_val_predict = mlr.predict(RF_val_x)
          #XG_val_predict = XG_model.predict(XG_val_x)

          RF_val_Pre = []
          for i in range(len(RF_val_predict)):
              Pred2 = RF_val_predict[i]
              RF_val_Pre.append(Pred2)
          #print('val_pre:', XG_val_Pre)


          val_mae = mean_absolute_error(val_y1, RF_val_Pre)
          #print('val_mae:', val_mae)
          val_rmse = sqrt(mean_squared_error(val_y1, RF_val_Pre))
          #print('val_mse:', val_mse)
          #val_R2 = r2_score(val_y1, val_Pred)

          i_VAL_MAE += val_mae
          i_VAL_RMSE += val_rmse



          va = []
          vb = abs(RF_val_Pre - val_y1)
          for i in range(len(vb)):
              if vb[i] <= 0.1:
                  va.append(vb[i])

          l = len(va)

          val_accuracy1 = float(l) / len(vb)
          # val_accuracy1 = ('%.2f' % val_accuracy1)

          val_accuracy2 = val_accuracy2 + float(val_accuracy1)
          val_Accuracy = val_accuracy2 / q2

          if val_Accuracy > val_bestacc:
              val_bestacc = val_Accuracy
              # best_model = model

          # print('valid_batch_Acc', val_Accuracy)

      if val_bestacc > epoch_acc:
          epoch_acc = val_bestacc
          epo = epoch

      i_VAL_MAE /= q2
      #print('i_VAL_MAE:', i_VAL_MAE)
      i_VAL_RMSE /= q2
      #print('i_VAL_MSE:', i_VAL_MSE)

      VAL_MAE += i_VAL_MAE
      VAL_RMSE += i_VAL_RMSE

  #print('all_epoch_mae', VAL_MAE)
  aver_mae = VAL_MAE / q1
  #print('epoch_mae:', aver_mae)  # epoch的mae
  MAE1.append(aver_mae)

  #print('all_epoch_mse', VAL_MSE)
  aver_rmse = VAL_RMSE / q1
  #print('epoch_mse:', aver_mse)  # epoch的mse
  RMSE1.append(aver_rmse)

  '''
  # print('train_Acc:', Accuracy)
  LOSS = []
  loss = np.array(loss_list)
  for i in range(len(loss)):
      loss1 = loss[i]
      LOSS.append(loss1)
   '''
  # print(LOSS)

  print('val_best_Acc:', val_bestacc)
  print('val_mae:', aver_mae)
  print('val_rmse:', aver_rmse)
  print("save model")

  torch.save(gcn_model.state_dict(), f'epoch_{epoch}_model.pth')
  #RF_model.get_booster().save_model(f'epoch_{epoch}_xgb.model')
  joblib.dump(mlr, f'epoch_{epoch}_mlr.pkl')

print('*************************************')
print('best_epo:', epo)
print('best_epoch_acc:', epoch_acc)
print('MAE:', MAE)
print('min_MAE_epoch', np.argmin(MAE))
print('min_MAE', np.min(MAE))
print('RMSE:', RMSE)

print('val_MAE:', MAE1)
print('min_val_MAE_epoch', np.argmin(MAE1))
print('min_val_MAE', np.min(MAE1))
print('val_RMSE:', RMSE1)
#print('loss：', LOSS)

import matplotlib.pyplot as plt

'''

plt.figure()
plt.subplot(221)
plt.plot(loss_list)
plt.xlabel('Num iteration')
plt.ylabel('Loss')
plt.legend()
plt.show()
'''

'''


num = q1 * EPOCH

plt.figure(figsize=(8, 3))

plt.subplot(121)
x = np.linspace(0, num, num=num)
plt.plot(x, LOSS, label='Loss')
plt.plot(MAE, label='mae')
plt.plot(RMSE, label='rmse')
plt.xlabel('Num iteration')
plt.ylabel('Train')
plt.legend()

#plt.figure()
plt.subplot(122)
# plt.ylim(0.2, 0.5)
plt.plot(MAE1, label='val_mae')
plt.plot(RMSE1, label='val_rmse')
plt.xlabel('Num iteration')
plt.ylabel('Valid')
plt.legend()

plt.show()

'''






'''
plt.scatter(x, val_Pred, marker='v', label='pred_y')
plt.scatter(x, y1, marker='o', label='y')

plt.xlabel('x')
plt.ylabel('y')
plt.legend(loc='lower right', fontsize=5, frameon=True, fancybox=True, framealpha=0.2, borderpad=0.3,
              ncol=1, markerfirst=True, markerscale=1, numpoints=1, handlelength=3.5)
plt.show()
'''