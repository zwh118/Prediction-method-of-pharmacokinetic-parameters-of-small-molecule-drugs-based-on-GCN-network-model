import openpyxl
from rdkit import Chem
from rdkit.Chem import Descriptors
from rdkit.Chem import Lipinski
from pprint import pprint
import xlrd

# Batch import of Excel data

file = xlrd.open_workbook(r'C:\Users\Desktop\oba1.xlsx')
ws = file.sheet_by_name('Sheet1')
datasets = []
c = 1
for r in range(1, ws.nrows):
    col = []
    datasets.append(ws.cell(r, c).value)
mols = []
for dataset in datasets:
    m = Chem.MolFromSmiles(dataset)# Read in smiles to obtain molecular objects
    mols.append(m)
#pprint(mols)

# Calculate molecular descriptors
i = len(mols)
print(i)

list = []
for o in range(0, i):
    print(o)
    n = mols[o]
    #print(n)
    tpsa_m = Descriptors.TPSA(n)# Topological Polarity Surface Area (TPSA)
    mol_weight = Descriptors.MolWt(n)# molecular weight
    H_Acc = Lipinski.NumHAcceptors(n)# Number of hydrogen bond acceptors
    H_Don = Lipinski.NumHDonors(n)# Hbond donor
    RTB = Lipinski.NumRotatableBonds(n)# Number of rotatable keys
    HAC = Lipinski.HeavyAtomCount(n)# Number of heavy atoms
    AR = Lipinski.NumAliphaticRings(n)# AR
    logp = Descriptors.MolLogP(n)# logp
    list1 = [tpsa_m, mol_weight, H_Acc, H_Don, RTB, HAC, AR, logp]
    list.append(list1)
pprint(list)

# Export data to Excel
wb = openpyxl.Workbook()
ws = wb.create_sheet('Sheet1')

heads = ['tpsa_m', 'mol_weight', 'H_Acc', 'H_Don', 'RTB', 'HAC', 'AR', 'logp']
l = len(list)
p = len(heads)

for r in range(0, l):
    t = r+1
    for c in range(0, p):
        p = c+1
        ws.cell(row=t, column=p).value = (list[r])[c]
        #print((list[r])[c])

wb.save('C:/Users/Desktop/s2.xlsx')
