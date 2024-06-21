# -*- coding: utf-8 -*-
"""
@author: wangying

@date:2021.09.17
"""
from pprint import pprint
from sklearn import model_selection as cv
import xlrd
import openpyxl
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

file = xlrd.open_workbook(r'C:\Users\Desktop\oba1.xlsx')
ws = file.sheet_by_name('0.3')

num = 133

y = []
x = []
for r in range(1, num+1):
    m = ws.cell(r, 10).value
    m = [m]
    y.append(m)

    x1 = []
    for c in range(0, 10):
        n = ws.cell(r, c).value

        x1.append(n)

    x.append(x1)
x_train_and_valid, x_test, y_train_and_valid, y_test = cv.train_test_split(x, y, test_size=0.2, random_state=10, shuffle=True)
x_train, x_valid, y_train, y_valid = cv.train_test_split(x_train_and_valid, y_train_and_valid, test_size=0.25, random_state=10, shuffle=True)

pprint(len(x_train))
pprint(len(x_valid))
pprint(len(x_test))
pprint(len(y_train))
pprint(len(y_valid))
pprint(len(y_test))


wb = openpyxl.Workbook()
def Write(x,y,w):
    ws = wb.create_sheet(w)
    l = len(x)
    p = 10
    for r in range(0, l):
        t = r + 1
        for c in range(0, p):
            p = c + 1
            ws.cell(row=t, column=p).value = (x[r])[c]
            ws.cell(row=t, column=11).value = (y[r])[0]


Write(x_train, y_train, 'train')
wb.save('C:/Users/Desktop/oba(0.3).xlsx')
Write(x_test, y_test, 'test')
wb.save('C:/Users/Desktop/oba(0.3).xlsx')
Write(x_valid, y_valid, 'valid')
wb.save('C:/Users/Desktop/oba(0.3).xlsx')

